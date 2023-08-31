
import json
import openpyxl

def read_from_file_new(file_name):
    
    with open(file_name, 'r') as file:
        _dataframe = openpyxl.load_workbook(file_name)
        dataframe = _dataframe.active
        datastr = ''
        stopExecution = False
        for row in range(1, dataframe.max_row):

            if (stopExecution == True):
                break
            
            adunitname = ''
            floorvalue = ''
            for col in dataframe.iter_cols(1, 2):
                
                if (str(col[row].value) == "None"):
                    datastr = datastr.removesuffix(', ')
                    stopExecution = True
                    break

                if (col[row].column_letter == 'A'):
                    adunitname = ' "'+ str(col[row].value) + '": '
                elif (col[row].column_letter == 'B'):
                    if (row != (dataframe.max_row-1)):
                        floorvalue = str(col[row].value) + ','
                    else:
                        floorvalue = str(col[row].value) + ''
                
            datastr = datastr + adunitname + floorvalue

        #print("datastr = ", datastr)
        _json_data = '{"configPattern": "_AU_", "config": { "default": { "floors": { "data": { "modelgroups": [ { "schema": {"fields": [ "adUnitCode"]},"values": {' + datastr + '},"modelversion": "model_1"}],"currency": "USD"}}}}}'
        json_data = json.loads(_json_data)
        #print(json_data)
        with open('data.json', 'w') as json_file:
            json.dump(json_data, json_file)


read_from_file_new('Book5.xlsx')


def read_from_file_old(file_name):
    
    with open(file_name, 'r') as file:
        _dataframe = openpyxl.load_workbook(file_name)
        dataframe = _dataframe.active
        datastr = ''
        stopExecution = False
        for row in range(1, dataframe.max_row):

            if (stopExecution == True):
                break
            
            adunitname = ''
            floorvalue = ''
            for col in dataframe.iter_cols(1, 2):
                
                if (str(col[row].value) == "None"):
                    datastr = datastr.removesuffix(', ')
                    stopExecution = True
                    break

                if (col[row].column_letter == 'A'):
                    adunitname = ' "'+ str(col[row].value) + '": {'
                elif (col[row].column_letter == 'B'):
                    if (row != (dataframe.max_row-1)):
                        floorvalue = ' "bidfloor": ' + str(col[row].value) + ', "bidfloorcur": "USD" }, '
                    else:
                        floorvalue = ' "bidfloor": ' + str(col[row].value) + ', "bidfloorcur": "USD" } '
                
            datastr = datastr + adunitname + floorvalue

        #print("datastr = ", datastr)
        _json_data = '{"configPattern": "_AU_", "config": {' + datastr + '} }'
        json_data = json.loads(_json_data)
        #print(json_data)
        with open('data.json', 'w') as json_file:
            json.dump(json_data, json_file)


#read_from_file_old('Book5.xlsx')